import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side

class ExcelReportGenerator:
    def __init__(self, dataframe):
        self.dataframe = dataframe
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def add_header(self):
        self.ws.merge_cells("A1:A3")
        self.ws.merge_cells("B1:B3")
        self.ws.merge_cells("C1:C3")
        self.ws["A1"] = "Название"
        self.ws["B1"] = "За указанный период"
        self.ws["C1"] = "За все время"
        alignment = Alignment(horizontal="center", vertical="center")
        self.ws["A1"].alignment = alignment
        self.ws["B1"].alignment = alignment
        self.ws["C1"].alignment = alignment
    
    def add_data(self):
        for r in dataframe_to_rows(self.dataframe, index=False, header=True):
            self.ws.append(r)
    
    def apply_borders_and_alignment(self):
        thin_border = Border(
            left=Side(style="thick"),
            right=Side(style="thick"),
            top=Side(style="thick"),
            bottom=Side(style="thick")
        )
        alignment = Alignment(horizontal="center", vertical="center")
        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = alignment
    
    def adjust_column_widths(self):
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter  # Получение буквы столбца
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column].width = adjusted_width
    
    def save_report(self, filename="report.xlsx"):
        self.wb.save(filename)

    def prepare_file(self):
        file_stream = io.BytesIO()
        self.wb.save(file_stream)
        file_stream.seek(0)
        return file_stream
